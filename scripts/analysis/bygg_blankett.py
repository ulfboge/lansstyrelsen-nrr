"""Bygger blankett_forvaltarkunskap_nnk.xlsx - insamlingsblankett for
forvaltarnas kunskap om livsmiljotyper, med kolumner som mappar mot
KartLitS granskningslager och NNK Ajourhalla.

Indata : data/nnk/nnk_d.json (fran nnk_kunskapslage.py), kodnamn.json,
         data/nnk/nnk_yta_med_sitecode.csv (fallback: docs/underlag/kartering.csv)
Utdata : docs/nnk/blankett_forvaltarkunskap_nnk.xlsx
"""
import csv
import json
import os
import collections
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

REPO = Path(__file__).resolve().parents[2]
NNK_DATA = REPO / "data" / "nnk"
NNK_DOCS = REPO / "docs" / "nnk"
UNDERLAG = REPO / "docs" / "underlag"

FONT = "Arial"
INK = "FF1F1F1F"
HDR_FILL = PatternFill("solid", fgColor="FF1F3864")
SUB_FILL = PatternFill("solid", fgColor="FFD9E2F3")
FYLL_FILL = PatternFill("solid", fgColor="FFFFFF99")   # gula = fylls i av forvaltaren
LAS_FILL = PatternFill("solid", fgColor="FFF2F2F2")    # gra = forifyllt, andra inte
P1_FILL = PatternFill("solid", fgColor="FFFCE4D6")
THIN = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MARIN = {"1110", "1130", "1140", "1150", "1152", "1153", "1154", "1160",
         "1170", "1174", "1000"}
LIMN = {"3100", "3110", "3130", "3150", "3160", "3200", "3210", "3212",
        "3213", "3214", "3220", "3260", "3261", "3265", "3266", "3000"}

TILLSTAND = ['"Gott,Icke gott,Okant - kan ej bedoma,Blandat - se andelar"']
JANEJ = ['"Ja,Nej,Vet ej"']
JUSTERING = ['"Inget behov av justering,Andring till annan livsmiljotyp,'
             'Andring till utvecklingsmark,Osaker - kan ej bedoma om livsmiljotyp eller inte,'
             'Obestamd - kan ej bedoma vilken livsmiljotyp"']
UTBREDNING = ['"Inget behov av justering,Yttergranser kvalitetsforbattring,'
              'Yttergranser andrad utbredning,Behov av att dela upp ytan flera livsmiljotyper"']
KONTROLL = ['"Typiska och karakteristiska arter,Strukturer,Havd,'
            'Funktioner (hydrologi storningar),Morfologi (jordart formationer),'
            'Annan negativ paverkan"']
METOD = ['"Faltbesok,Faltinventering (standardiserad metodik),'
         'Skrivbord / Granska mot andra underlag,Annan metod"']
GRUND = ['"Eget faltbesok,Standardiserad inventering (uppfoljning/AoB),'
         'Skotselplan eller bevarandeplan,Konsultrapport eller PM,'
         'Betesstod / jordbruksblock,Allman lokalkannedom,Ingen grund - osaker"']
HAVD = ['"Aktiv havd pagar,Havd har upphort,Havd men otillracklig,'
        'Ej havdberoende,Vet ej"']


def kolumn(ws, col, bredd):
    ws.column_dimensions[get_column_letter(col)].width = bredd


def rubrikrad(ws, rad, kolumner, forifyllt_tom):
    """Skriv rubrikrad. forifyllt_tom = index (1-baserat) dar de gula borjar."""
    for i, (txt, bredd, hjalp) in enumerate(kolumner, start=1):
        c = ws.cell(row=rad, column=i, value=txt)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BOX
        if hjalp:
            c.comment = Comment(hjalp, "Metodik NNK")
        kolumn(ws, i, bredd)
    ws.row_dimensions[rad].height = 46


def las_data():
    d = json.load(open(NNK_DATA / "nnk_d.json", encoding="utf-8"))
    namn = json.load(open(Path(__file__).with_name("kodnamn.json"), encoding="utf-8"))
    rader = []
    for o in d["objekt"]:
        for kod, areal in sorted(o["koder"].items(), key=lambda x: -x[1]):
            if kod in MARIN or kod in LIMN or areal < 0.05:
                continue
            info = namn.get(kod, {})
            rader.append({
                "tier": o["tier"],
                "sitecode": o["sitecode"],
                "objekt": o["namn"],
                "kod": kod,
                "namn": info.get("namn", ""),
                "under": info.get("under", ""),
                "kategori": info.get("kat", ""),
                "areal": round(areal, 2),
                "havdber": kod in {
                    "1630", "1631", "5130", "5133", "6110", "6210", "6230",
                    "6270", "6280", "6410", "6412", "6430", "6510", "6520",
                    "8230", "8231", "8232", "9070", "9071", "9072"},
                "score": o["score"],
            })
    ordning = {"P1": 0, "P2": 1, "P3": 2, "P4": 3}
    rader.sort(key=lambda r: (ordning[r["tier"]], -r["score"], -r["areal"]))
    return rader, namn


def las_atgardas():
    """Atgardas-ytor per objekt och livsmiljotyp.

    Kraver nnk_yta_med_sitecode.csv fran koppla_omraden.py. Utan den faller
    skriptet tillbaka pa kartering.csv och grupperar bara per livsmiljotyp.
    """
    kalla = NNK_DATA / "nnk_yta_med_sitecode.csv"
    if not os.path.exists(kalla):
        kalla = UNDERLAG / "kartering.csv"
    rows = list(csv.DictReader(open(kalla, encoding="utf-8-sig"), delimiter=";"))
    atg = [r for r in rows if r["KARTERINGS"].startswith("5")]
    har_omrade = "SITECODE" in rows[0]

    grupp = {}
    for r in atg:
        sc = (r.get("SITECODE") or "").strip() or "—"
        omr = (r.get("OMRADE") or "").strip() or "Ej kopplat till Natura 2000"
        nt = r["NATURTYP"].strip()
        ha = float(r.get("overlapp_ha") or 0)
        y = (r["REDIGERATG"] or "")[:4] or (r["SKAPATDATU"] or "")[:4]
        nyckel = (sc, omr, nt)
        d = grupp.setdefault(nyckel, {"antal": 0, "ha": 0.0, "ar": set()})
        d["antal"] += 1
        d["ha"] += ha
        if y:
            d["ar"].add(y)
    return atg, grupp, har_omrade


# ---------------------------------------------------------------- byggnad
wb = openpyxl.Workbook()

# ============ Flik 1: Lasanvisning ============
ws = wb.active
ws.title = "Läsanvisning"
ws.sheet_view.showGridLines = False
kolumn(ws, 1, 3); kolumn(ws, 2, 30); kolumn(ws, 3, 96)

rader_text = [
    ("H1", "Blankett — förvaltarnas kunskap om livsmiljötyper", ""),
    ("P", "Länsstyrelsen i Södermanlands län · Naturskyddsenheten · NNK/NRF ref. 2451-2026 · version 1.1, 2026-08-17", ""),
    ("", "", ""),
    ("H2", "Vad blanketten är till för", ""),
    ("P", "Mycket av det länsstyrelsen vet om sina skyddade områden finns hos reservatsförvaltarna på "
          "Naturvårdsenheten och inte i Natura naturtypskartan (NNK). Den här blanketten samlar in den "
          "kunskapen i en form som går att föra rakt in i KartLitS granskningslager och därefter i NNK.", ""),
    ("P", "Naturvårdsverket godkänner uttryckligen lokalkännedom som kunskapskälla — se Handledning för NNK "
          "avsnitt 4.1 och lathunden för WebbGIS-granskning.", ""),
    ("", "", ""),
    ("H2", "Så här använder du den", ""),
    ("LI", "1.", "Öppna fliken Blankett. Filtrera kolumn A–C till de objekt en viss förvaltare ansvarar för "
                 "och skicka den delen — ingen behöver se alla 628 raderna."),
    ("LI", "2.", "GRÅ kolumner (A–I) är förifyllda ur NNK-statistiken. Ändra dem inte."),
    ("LI", "3.", "GULA kolumner (J–W) fylls i av förvaltaren. Alla har rullistor utom fritextfälten."),
    ("LI", "4.", "Rad 3 är ett ifyllt exempel. Radera den innan blanketten skickas ut, eller låt den stå kvar "
                 "som förlaga — men blanda inte ihop den med riktiga svar."),
    ("LI", "5.", "Inget fält är obligatoriskt. Ett tomt fält betyder \"vet ej\" och är ett giltigt svar. "
                 "En gissning är det inte."),
    ("LI", "6.", "Fliken Åtgärdas-ytor är den bästa öppningsfrågan i ett förvaltarsamtal — 141 ytor där "
                 "basinventeringen inte kunde bestämma naturtypen, koncentrerade till sju objekt."),
    ("", "", ""),
    ("H2", "Tre regler att ha med sig i samtalet", ""),
    ("LI", "R1", "En igenvuxen äng är fortfarande en äng. Om orsaken är utebliven skötsel ska livsmiljötypen "
                 "stå kvar och tillståndet sättas till Icke gott — inte klassas om till annan naturtyp. "
                 "(Lathunden, \"Vad kan vi ändra på?\")"),
    ("LI", "R2", "Kunskap som funnits men aldrig registrerats är en Komplettering, inte en Faktisk förändring. "
                 "Det avgör om länet rapporterar in en arealförändring som aldrig hänt. "
                 "(Handledningen 5.4)"),
    ("LI", "R3", "Fråga alltid efter årtal. Utan datering går FAQ fråga 4:s krav på \"hur aktuell bedömningen är\" "
                 "inte att besvara."),
    ("", "", ""),
    ("H2", "Flikar", ""),
    ("LI", "Blankett", "628 rader, en per objekt × livsmiljötyp. Sorterad med prioritet 1 först."),
    ("LI", "Åtgärdas-ytor", "De 141 ytor med karteringsstatus 5, per objekt och livsmiljötyp."),
    ("LI", "Kodlistor", "Tillåtna värden. Källa till rullistorna — redigera inte."),
    ("LI", "Fältmappning", "Vilken blankettkolumn som hamnar i vilket fält i granskningslagret och NNK."),
    ("", "", ""),
    ("H2", "Källor", ""),
    ("P", "Förifylld areal och objektsurval: NNK-statistik per Natura 2000-område D-län, uttag 2026-01-20 "
          "(Naturvårdsverket). Åtgärdas-ytorna: NNK-lagrets attributtabell, uttag från ArcGIS Pro 2026-08-17. "
          "Kodvärden: Handledning för NNK 2026-07-03 samt Lathund granskning WebbGIS-KartLitS 2026-07-10.", ""),
]

r = 2
for typ, a, b in rader_text:
    if typ == "H1":
        c = ws.cell(row=r, column=2, value=a); c.font = Font(name=FONT, size=16, bold=True, color=INK)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 24
    elif typ == "H2":
        c = ws.cell(row=r, column=2, value=a); c.font = Font(name=FONT, size=11, bold=True, color="FF1F3864")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    elif typ == "P":
        c = ws.cell(row=r, column=2, value=a); c.font = Font(name=FONT, size=10, color=INK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 15 * (1 + len(a) // 150)
    elif typ == "LI":
        c = ws.cell(row=r, column=2, value=a); c.font = Font(name=FONT, size=10, bold=True, color=INK)
        c.alignment = Alignment(vertical="top")
        c2 = ws.cell(row=r, column=3, value=b); c2.font = Font(name=FONT, size=10, color=INK)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 14 * (1 + len(b) // 105)
    r += 1

# ============ Flik 2: Kodlistor ============
kl = wb.create_sheet("Kodlistor")
kl.sheet_view.showGridLines = False
listor = [
    ("Tillstånd", ["Gott", "Icke gott", "Okänt – kan ej bedöma", "Blandat – se andelar"]),
    ("Hävdstatus", ["Aktiv hävd pågår", "Hävd har upphört", "Hävd men otillräcklig",
                    "Ej hävdberoende", "Vet ej"]),
    ("Grund för bedömning", ["Eget fältbesök", "Standardiserad inventering (uppföljning/ÄoB)",
                             "Skötselplan eller bevarandeplan", "Konsultrapport eller PM",
                             "Betesstöd / jordbruksblock", "Allmän lokalkännedom",
                             "Ingen grund – osäker"]),
    ("Livsmiljötyp, behov av justering", ["Inget behov av justering", "Ändring till annan livsmiljötyp",
                                          "Ändring till utvecklingsmark",
                                          "Osäker – kan ej bedöma om livsmiljötyp eller inte",
                                          "Obestämd – kan ej bedöma vilken livsmiljötyp"]),
    ("Utbredning, behov av justering", ["Inget behov av justering", "Yttergränser, kvalitetsförbättring",
                                        "Yttergränser, ändrad utbredning",
                                        "Behov av att dela upp ytan, flera livsmiljötyper"]),
    ("Vad ska kontrolleras", ["Typiska och karakteristiska arter", "Strukturer", "Hävd",
                              "Funktioner (hydrologi, störningar)",
                              "Morfologi (jordart, formationer)", "Annan negativ påverkan"]),
    ("Metod för kontroll", ["Fältbesök", "Fältinventering (standardiserad metodik)",
                            "Skrivbord / Granska mot andra underlag", "Annan metod"]),
    ("Ja/Nej", ["Ja", "Nej", "Vet ej"]),
]
c = kl.cell(row=1, column=1, value="Kodlistor – källa till rullistorna i Blankett. Redigera inte.")
c.font = Font(name=FONT, size=11, bold=True, color=INK)
kl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
for i, (rub, varden) in enumerate(listor, start=1):
    h = kl.cell(row=3, column=i, value=rub)
    h.font = Font(name=FONT, size=9, bold=True, color="FFFFFFFF")
    h.fill = HDR_FILL
    h.alignment = Alignment(wrap_text=True, vertical="center")
    kolumn(kl, i, 30)
    for j, v in enumerate(varden, start=4):
        cc = kl.cell(row=j, column=i, value=v)
        cc.font = Font(name=FONT, size=9, color=INK)
kl.row_dimensions[3].height = 30

def omr(kolnr, antal):
    L = get_column_letter(kolnr)
    return f"Kodlistor!${L}$4:${L}${3+antal}"

# ============ Flik 3: Blankett ============
bl = wb.create_sheet("Blankett", 1)
bl.sheet_view.showGridLines = False
rader, namn = las_data()

KOL = [
    ("Prio", 6, "P1–P4 enligt prioriteringen i arbetsplanen. P1 = hävdberoende eller sällsynt livsmiljötyp."),
    ("Sitecode", 12, "Natura 2000-områdets kod."),
    ("Områdesnamn", 26, None),
    ("Kod", 7, "NNK naturtypskod / livsmiljötyp."),
    ("Livsmiljötyp", 30, None),
    ("Kategori", 11, None),
    ("Areal (ha)", 10, "Karterad areal av denna livsmiljötyp inom objektet, enligt NNK-uttag 2026-01-20."),
    ("Hävdberoende", 12, "Sätts automatiskt utifrån livsmiljötypen. Hävdberoende marker har högsta prioritet enligt FAQ fråga 11."),
    ("Förvaltare", 16, "Fyll i vem som ansvarar för objektet, så går blanketten att dela ut."),
    # --- gula fran och med har (kolumn 10) ---
    ("Tillstånd", 15, "Gott / Icke gott / Okänt. OBS: en igenvuxen mark där orsaken är utebliven skötsel är Icke gott – inte en annan naturtyp."),
    ("Andel gott (%)", 10, "Fyll bara i om tillståndet varierar inom ytan. Summan av de tre andelarna bör bli 100."),
    ("Andel ej gott (%)", 11, "Se ovan."),
    ("Andel osäker (%)", 11, "Se ovan."),
    ("Hävdstatus", 17, "Gäller hävdberoende marker. Avgörande för tillståndsbedömningen."),
    ("Grund för bedömning", 24, "Vad vet du det på? Går till Kommentar_metod i granskningslagret och KOMMENTAR i NNK."),
    ("År för senaste bedömning", 12, "Årtal, t.ex. 2024. Ett osäkert årtal är bättre än inget. Går till habitat_period_lastdata_end."),
    ("Bedömare", 16, "Namn på den som gjort bedömningen – förvaltarens namn, inte handläggarens. Går till faltinventerare."),
    ("Livsmiljötyp, behov av justering", 26, "Välj bara om du anser att livsmiljötypen är fel eller osäker."),
    ("Föreslagen livsmiljötyp", 18, "Kod, t.ex. 9070. Vid utvecklingsmark blir detta målnaturtyp."),
    ("Utbredning, behov av justering", 25, "Gäller gränser och areal. Justeringar under minsta karteringsenhet (0,25 ha; 1 ha skog/våtmark) behöver inte göras."),
    ("Vad ska kontrolleras", 24, "Vad behöver kontrolleras för att fastställa livsmiljötyp eller tillstånd?"),
    ("Metod för kontroll", 22, "Framåtsyftande – vilken metod BÖR användas, inte hur du hittills gjort."),
    ("Kommentar", 40, "Fritext. Skriv gärna vad du är osäker på och vilka dokument som finns."),
]
GUL_FRAN = 10

rubrikrad(bl, 1, KOL, GUL_FRAN)
bl.freeze_panes = "J2"
bl.auto_filter.ref = f"A1:{get_column_letter(len(KOL))}{len(rader)+2}"

# Exempelrad
ex = ["P1", "SE0220118", "Labro ängar", "6270", "Silikatgräsmarker", "Gräsmark", 21.4, "Ja", "(exempel)",
      "Blandat – se andelar", 60, 40, 0, "Hävd men otillräcklig",
      "Eget fältbesök", 2024, "Exempel Exempelsson",
      "Inget behov av justering", "", "Inget behov av justering",
      "Hävd", "Fältbesök",
      "Norra delen betas fortfarande, södra har vuxit igen med sly sedan djuren minskade 2021. "
      "Skötselplan från 2016 finns i pärm på enheten."]
for i, v in enumerate(ex, start=1):
    c = bl.cell(row=2, column=i, value=v)
    c.font = Font(name=FONT, size=9, italic=True, color="FF808080")
    c.fill = LAS_FILL if i < GUL_FRAN else PatternFill("solid", fgColor="FFFFF7D6")
    c.border = BOX
    c.alignment = Alignment(wrap_text=(i == len(KOL)), vertical="top")
bl.cell(row=2, column=1).comment = Comment(
    "EXEMPELRAD – visar förväntat format. Radera raden innan blanketten skickas ut, "
    "eller låt den stå kvar som förlaga.", "Metodik NNK")
bl.row_dimensions[2].height = 30

for j, rd in enumerate(rader, start=3):
    vals = [rd["tier"], rd["sitecode"], rd["objekt"], rd["kod"],
            (rd["namn"] + (" – " + rd["under"] if rd["under"] else "")),
            rd["kategori"], rd["areal"], "Ja" if rd["havdber"] else "Nej", None]
    for i, v in enumerate(vals, start=1):
        c = bl.cell(row=j, column=i, value=v)
        c.font = Font(name=FONT, size=9, color=INK)
        c.fill = P1_FILL if (i == 1 and rd["tier"] == "P1") else LAS_FILL
        c.border = BOX
        if i == 7:
            c.number_format = "#,##0.0"
    for i in range(GUL_FRAN, len(KOL) + 1):
        c = bl.cell(row=j, column=i)
        c.font = Font(name=FONT, size=9, color=INK)
        c.fill = FYLL_FILL
        c.border = BOX
        if i in (11, 12, 13, 16):
            c.number_format = "0"

sista = len(rader) + 2


def dv(kolnr, kodkol, antal, tillat_tom=True):
    d = DataValidation(type="list", formula1=omr(kodkol, antal), allow_blank=tillat_tom)
    bl.add_data_validation(d)
    L = get_column_letter(kolnr)
    d.add(f"{L}3:{L}{sista}")


dv(10, 1, 4)    # Tillstand
dv(14, 2, 5)    # Havdstatus
dv(15, 3, 7)    # Grund
dv(18, 4, 5)    # Livsmiljotyp justering
dv(20, 5, 4)    # Utbredning
dv(21, 6, 6)    # Vad ska kontrolleras
dv(22, 7, 4)    # Metod

for kolnr, mini, maxi, txt in [(11, 0, 100, "Andel i procent, 0–100."),
                               (12, 0, 100, "Andel i procent, 0–100."),
                               (13, 0, 100, "Andel i procent, 0–100."),
                               (16, 1990, 2027, "Årtal mellan 1990 och 2027.")]:
    d = DataValidation(type="whole", operator="between", formula1=mini, formula2=maxi,
                       allow_blank=True, showErrorMessage=True,
                       errorTitle="Ogiltigt värde", error=txt)
    bl.add_data_validation(d)
    L = get_column_letter(kolnr)
    d.add(f"{L}3:{L}{sista}")

# ============ Flik 4: Atgardas-ytor ============
at = wb.create_sheet("Åtgärdas-ytor")
at.sheet_view.showGridLines = False
atg, grupp, har_omrade = las_atgardas()
kolumn(at, 1, 3); kolumn(at, 2, 12); kolumn(at, 3, 26); kolumn(at, 4, 8)
kolumn(at, 5, 48); kolumn(at, 6, 9); kolumn(at, 7, 10); kolumn(at, 8, 15); kolumn(at, 9, 44)

c = at.cell(row=2, column=2, value="Åtgärdas-ytor — basinventeringens öppna kunskapsluckor")
c.font = Font(name=FONT, size=14, bold=True, color=INK)
at.merge_cells(start_row=2, start_column=2, end_row=2, end_column=9)
txt = (f"{len(atg)} polygoner i länet har karteringsstatus 5 – Åtgärdas. Den publika produktbeskrivningen "
       "kallar koden \"äldre kod som finns kvar från basinventeringen\" och den betydde att det behövdes "
       "kompletterande uppgifter för att bestämma naturtypen. Datat bekräftar det: 139 av 141 har ursprung "
       "BIDOS, samtliga redigerades 2007–2008 (tre stycken 2019) och alla har naturtypsstatus 5 Ej bedömd. "
       "Det är alltså en dokumenterad kunskapslucka som stått öppen i nitton år — och den ligger nästan "
       "uteslutande i hävdberoende marker, den kategori FAQ fråga 11 sätter högst. "
       "Öppningsfråga i förvaltarsamtalet: \"basinventeringen kunde inte bestämma naturtypen här — vet du vad det är?\"")
c = at.cell(row=3, column=2, value=txt)
c.font = Font(name=FONT, size=10, color=INK)
c.alignment = Alignment(wrap_text=True, vertical="top")
at.merge_cells(start_row=3, start_column=2, end_row=6, end_column=9)

for i, h in enumerate(["Sitecode", "Områdesnamn", "Kod", "Livsmiljötyp", "Antal ytor",
                       "Areal (ha)", "Senast redigerad", "Anteckning från förvaltarsamtal"], start=2):
    c = at.cell(row=8, column=i, value=h)
    c.font = Font(name=FONT, size=9, bold=True, color="FFFFFFFF")
    c.fill = HDR_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BOX
at.row_dimensions[8].height = 28

HAVDKODER = {"1630", "1631", "5130", "5133", "6110", "6210", "6230", "6270",
             "6280", "6410", "6412", "6430", "6510", "6520", "8230", "8231",
             "8232", "9070", "9071", "9072"}
r = 9
poster = sorted(grupp.items(), key=lambda x: (-x[1]["antal"], x[0][0]))
# gruppera per objekt, storsta objekt forst
per_obj = {}
for (sc, omr, nt), v in grupp.items():
    per_obj.setdefault((sc, omr), 0)
    per_obj[(sc, omr)] += v["antal"]
ordn = sorted(per_obj.items(), key=lambda x: -x[1])
for (sc, omr), _tot in ordn:
    rader_obj = sorted([(nt, v) for (s2, o2, nt), v in grupp.items() if (s2, o2) == (sc, omr)],
                       key=lambda x: -x[1]["antal"])
    for nt, v in rader_obj:
        kod = nt.split(" - ")[0].strip()
        beskr = nt.split(" - ", 1)[1] if " - " in nt else nt
        arlist = ", ".join(sorted(y for y in v["ar"] if y))
        vals = [sc, omr, kod, beskr, v["antal"], round(v["ha"], 2), arlist, None]
        for i, val in enumerate(vals, start=2):
            c = at.cell(row=r, column=i, value=val)
            c.font = Font(name=FONT, size=9, color=INK)
            c.border = BOX
            c.fill = FYLL_FILL if i == 9 else (
                P1_FILL if (i == 4 and kod in HAVDKODER) else LAS_FILL)
            if i in (6, 7):
                c.alignment = Alignment(horizontal="right")
            if i == 7:
                c.number_format = "#,##0.00"
        r += 1
c = at.cell(row=r, column=3, value="Summa")
c.font = Font(name=FONT, size=9, bold=True, color=INK)
for kol, formel in ((6, f"=SUM(F9:F{r-1})"), (7, f"=SUM(G9:G{r-1})")):
    c = at.cell(row=r, column=kol, value=formel)
    c.font = Font(name=FONT, size=9, bold=True, color=INK)
    c.alignment = Alignment(horizontal="right")
    c.border = BOX
    if kol == 7:
        c.number_format = "#,##0.00"
at.auto_filter.ref = f"B8:I{r-1}"

kallnot = ("Källa: NNK_YTA ur Naturtypskartan_D (publik Natura naturtypskarta), filtrerat på "
           "karteringsstatus 5 – Åtgärdas. Områdestillhörighet framtagen med "
           "scripts/analysis/koppla_omraden.py: överlappsbaserad koppling mot det rikstäckande "
           "SCI-lagret från Naturvårdsregistret. Samtliga 141 ytor ligger inom Natura 2000. "
           "Orange markering i kolumnen Livsmiljötyp = hävdberoende livsmiljötyp.")
c = at.cell(row=r + 2, column=2, value=kallnot)
c.font = Font(name=FONT, size=9, italic=True, color="FF808080")
c.alignment = Alignment(wrap_text=True, vertical="top")
at.merge_cells(start_row=r + 2, start_column=2, end_row=r + 4, end_column=9)

# ============ Flik 5: Faltmappning ============
fm = wb.create_sheet("Fältmappning")
fm.sheet_view.showGridLines = False
kolumn(fm, 1, 3); kolumn(fm, 2, 32); kolumn(fm, 3, 34); kolumn(fm, 4, 26); kolumn(fm, 5, 46)
c = fm.cell(row=2, column=2, value="Från blankett till granskningslager till NNK")
c.font = Font(name=FONT, size=14, bold=True, color=INK)
fm.merge_cells(start_row=2, start_column=2, end_row=2, end_column=5)
c = fm.cell(row=3, column=2,
            value="Registrera först i KartLitS granskningslager (förslagsnivå), därefter i NNK Ajourhålla "
                  "efter avstämning. Avvakta med tillståndsfälten i NNK tills de nya attributen driftsatts "
                  "i slutet av september 2026 (FAQ fråga 30).")
c.font = Font(name=FONT, size=10, color=INK)
c.alignment = Alignment(wrap_text=True, vertical="top")
fm.merge_cells(start_row=3, start_column=2, end_row=4, end_column=5)

for i, h in enumerate(["Blankettkolumn", "Granskningslager (WebbGIS)", "NNK Ajourhålla", "Att tänka på"], start=2):
    c = fm.cell(row=6, column=i, value=h)
    c.font = Font(name=FONT, size=9, bold=True, color="FFFFFFFF")
    c.fill = HDR_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BOX
fm.row_dimensions[6].height = 28

MAPP = [
    ("Tillstånd", "tillstand", "NATURTYPSSTATUS (1 Fullgod / 2 Icke fullgod / 5 Ej bedömd)",
     "Igenväxning på grund av utebliven skötsel ger 2 Icke fullgod — inte byte av naturtyp."),
    ("Andel gott / ej gott / osäker", "procent_gott, procent_ej_gott, procent_osaker",
     "Nya tillståndsfält", "Driftsätts i NNK slutet av september 2026. Ytan behöver då inte längre delas."),
    ("Hävdstatus", "kommentar_tillstand (fritext)", "KOMMENTAR",
     "Ingen egen kodlista i NNK — skriv i klartext, den är grunden för tillståndsbedömningen."),
    ("Grund för bedömning", "Kommentar_metod", "KOMMENTAR",
     "FAQ fråga 4 kräver att grunden anges. Fältet syns inte i den publika NNK — kontrollera mot NNK Ajourhålla."),
    ("År för senaste bedömning", "habitat_period_lastdata_end", "Slutdatum senaste inventering",
     "Nytt fält, tomt idag. Uppdatera framför allt slutdatum när du granskar."),
    ("Bedömare", "faltinventerare", "KOMMENTAR (namn + roll)",
     "Ange förvaltarens namn, inte handläggarens."),
    ("Livsmiljötyp, behov av justering", "justering", "NATURTYP",
     "Utpekade livsmiljötyper och prioriterade bevarandevärden ändras bara vid uppenbart fel (FAQ fråga 19)."),
    ("Föreslagen livsmiljötyp", "livsmiljötyp1–3", "NATURTYP eller MALNATURTYP1–3",
     "Vid utvecklingsmark blir de målnaturtyper — då måste NATURTYP vara en icke-natura-kod."),
    ("Utbredning, behov av justering", "utbredning", "Geometri",
     "Justera bara där redigeringen påverkar arealen. Minsta karteringsenhet: 0,25 ha generellt, "
     "1 ha skog och våtmark, 0,5 ha ädellöv/lövsump/svämskog, 2 ha ovan trädgränsen."),
    ("Vad ska kontrolleras", "kontroll1–3, kommentar_kontroll", "KARTERINGSSTATUS 5 (Åtgärdas) + KOMMENTAR",
     "Tre likadana rullistor i WebbGIS eftersom flerval saknas."),
    ("Metod för kontroll", "metod, Kommentar_metod", "—",
     "Framåtsyftande: vilken metod som bör användas, inte hur du hittills gjort."),
    ("Kommentar", "kommentar_livsmil_utbred / kommentar_tillstand", "KOMMENTAR",
     "Lämna aldrig tomt efter en redigering."),
    ("(sätts av dig, ej i blanketten)", "—", "KARTERINGSSTATUS",
     "2 Granskad vid skrivbordet för förvaltarkunskap och dokument. 3 Besökt i fält om förvaltaren "
     "faktiskt varit där nyligen. 4 Inventerad i fält endast vid standardiserad metodik."),
    ("(sätts av dig, ej i blanketten)", "—", "FÖRÄNDRINGSORSAK",
     "3 Komplettering när kunskapen fanns men aldrig registrerats. 1 Rättning vid felaktig kartering. "
     "2 Faktisk förändring endast när naturen faktiskt förändrats."),
]
r = 7
for a, b, c_, d_ in MAPP:
    for i, v in enumerate([a, b, c_, d_], start=2):
        cc = fm.cell(row=r, column=i, value=v)
        cc.font = Font(name=FONT, size=9, color=INK,
                       italic=a.startswith("(sätts"))
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        cc.border = BOX
        cc.fill = SUB_FILL if a.startswith("(sätts") else LAS_FILL
    fm.row_dimensions[r].height = 30
    r += 1

# --- utskriftsinstallningar ---
for s in wb.worksheets:
    s.sheet_properties.tabColor = "FF1F3864"
    s.page_setup.fitToWidth = 1
    s.page_setup.fitToHeight = 0
    s.sheet_properties.pageSetUpPr.fitToPage = True
    s.page_margins.left = s.page_margins.right = 0.4
    s.page_margins.top = s.page_margins.bottom = 0.5
    s.oddFooter.right.text = "&P / &N"
    s.oddFooter.left.text = "Blankett forvaltarkunskap NNK - Lansstyrelsen Sodermanland"
    s.oddFooter.left.size = 8
    s.oddFooter.right.size = 8

for s in (bl, at, fm):
    s.page_setup.orientation = "landscape"
# 23 kolumner blir oläsligt pa A4 - blanketten far A3 vid utskrift
bl.page_setup.paperSize = bl.PAPERSIZE_A3
bl.print_title_rows = "1:1"
at.print_title_rows = "8:8"
fm.print_title_rows = "6:6"
ws.page_setup.orientation = "portrait"
ws.print_area = f"B1:C{r}"

utfil = NNK_DOCS / "blankett_forvaltarkunskap_nnk.xlsx"
wb.save(utfil)
print(f"Skrev {utfil}: {len(rader)} blankettrader, "
      f"{len(grupp)} rader i Åtgärdas-listan ({len(atg)} ytor)")
